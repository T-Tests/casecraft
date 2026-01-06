import json
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.schema import TestSuite
from core.output import OutputFormat


def _join_lines(items: List[str]) -> str:
    return "\n".join(items) if items else ""


def _export_excel(suite: TestSuite, output_path: str) -> None:
    workbook = Workbook()

    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet(title="Test Cases")
    else:
        sheet.title = "Test Cases"

    headers = [
        "Use Case",
        "Test Case",
        "Preconditions",
        "Test Data",
        "Steps",
        "Priority",
        "Tags",
        "Expected Results",
        "Actual Results",
    ]

    sheet.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.alignment = Alignment(wrap_text=True)
        sheet.column_dimensions[get_column_letter(col_idx)].width = 30

    for test_case in suite.test_cases:
        sheet.append([
            test_case.use_case,
            test_case.test_case,
            _join_lines(test_case.preconditions),
            "\n".join(f"{k}: {v}" for k, v in test_case.test_data.items()),
            _join_lines(test_case.steps),
            test_case.priority,
            ", ".join(test_case.tags),
            _join_lines(test_case.expected_results),
            _join_lines(test_case.actual_results or []),
        ])

    # no need to delete default sheet when using workbook.active

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _export_json(suite: TestSuite, output_path: str) -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(suite.model_dump_json(indent=2))


def export(
    suite: TestSuite,
    output_format: OutputFormat,
    output_path: str,
) -> None:
    if output_format == OutputFormat.excel:
        _export_excel(suite, output_path)
    elif output_format == OutputFormat.json:
        _export_json(suite, output_path)
    else:
        raise ValueError(f"Unsupported output format: {output_format}")
